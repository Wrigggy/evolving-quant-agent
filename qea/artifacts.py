"""Artifact handling for the B-pile worker: extract the worker's openpyxl code,
render a produced .xlsx to faithful text, and orchestrate produce->capture->render.

The text rendering is the interim grader bridge (sub-project 1): it makes artifact
rubric criteria creditable by the existing text SoftJudge. NOTE (known v0.1 limit):
openpyxl-written formulas carry no computed value, so we render the formula STRING +
literal values, not formula results — value computation is sub-project 3."""
from __future__ import annotations

import shutil
from pathlib import Path


def extract_openpyxl_code(text: str) -> str | None:
    """Return the first ```python code block that builds + saves a workbook
    (mentions openpyxl AND saves), else None (text deliverable)."""
    if "```" not in text:
        return None
    parts = text.split("```")
    for i in range(1, len(parts), 2):           # odd indices are fenced blocks
        block = parts[i]
        # drop a leading fence language tag line (```python / ```py / ```Python …)
        first, _, rest = block.partition("\n")
        if first.strip().lower() in ("python", "py", "python3"):
            block = rest
        # tolerate whitespace inside the save call (.save( "x.xlsx" ))
        if "openpyxl" in block and ".save(" in block.replace(" ", ""):
            return block.strip()
    return None


def _strip_code_blocks(text: str) -> str:
    """Drop fenced ``` blocks, leaving the narrative."""
    if "```" not in text:
        return text.strip()
    parts = text.split("```")
    return "".join(parts[i] for i in range(0, len(parts), 2)).strip()


def render_xlsx(path, max_rows: int = 100, max_cols: int = 30) -> str:
    """Faithful text dump of a workbook: filename + per-sheet name/dims + a bounded
    grid of non-empty cells (literal values, and formula strings for formula cells)."""
    import openpyxl  # lazy: keeps qea.artifacts importable without the [artifacts] extra
    wb = openpyxl.load_workbook(path, data_only=False)
    lines = [f"[ARTIFACT FILE: {Path(path).name}]"]
    for ws in wb.worksheets:
        lines.append(f'Sheet "{ws.title}" ({ws.max_row}x{ws.max_column}):')
        for row in ws.iter_rows(max_row=min(ws.max_row, max_rows),
                                max_col=min(ws.max_column, max_cols)):
            for cell in row:
                if cell.value is not None:
                    lines.append(f"  {cell.coordinate}: {cell.value!r}")
        if ws.max_row > max_rows or ws.max_column > max_cols:
            lines.append(f"  ...(truncated to {max_rows}x{max_cols})")
    return "\n".join(lines)


def assemble_artifact_deliverable(llm_text: str, task, artifact_dir) -> str:
    """If the worker emitted openpyxl code, run it in the subprocess sandbox; on
    success persist the .xlsx under <artifact_dir>/<task_id>/ and return
    narrative + rendered artifact. Otherwise (text task or failed exec) return the
    text unchanged — never raises, so a bad workbook just degrades to its narrative."""
    code = extract_openpyxl_code(llm_text)
    if code is None:
        return llm_text
    from .sandbox import exec_artifact          # local import: sandbox is stdlib-only
    res = exec_artifact(code)
    if res.status != "success" or not res.paths:
        if res.work_dir:
            shutil.rmtree(res.work_dir, ignore_errors=True)
        return llm_text
    renderings = []
    for p in res.paths:
        renderings.append(render_xlsx(p))
        if artifact_dir is not None:
            dest = Path(artifact_dir) / task.task_id
            dest.mkdir(parents=True, exist_ok=True)
            shutil.copy2(p, dest / p.name)
    shutil.rmtree(res.work_dir, ignore_errors=True)
    narrative = _strip_code_blocks(llm_text)
    return (narrative + "\n\n" + "\n\n".join(renderings)).strip()
