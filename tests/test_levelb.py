"""Phase-4 Level-B loop: unit tests for the pure logic + a gated NexAU smoke test.

The NexAU-real pieces (run_worker / run_evolve_agent / the full loop) require an
API key + proxy and are gated behind QEA_LEVELB_SMOKE=1; everything else is
offline and deterministic.
"""
import os
from pathlib import Path

import pytest


@pytest.fixture(autouse=True)
def _dummy_llm_env(monkeypatch):
    """AgentConfig.from_yaml eagerly resolves ${env.LLM_*}; offline tests only need
    the config to PARSE, so provide harmless dummies (no network is made)."""
    monkeypatch.setenv("LLM_MODEL", "deepseek/deepseek-v4-pro")
    monkeypatch.setenv("LLM_BASE_URL", "https://openrouter.ai/api/v1")
    monkeypatch.setenv("LLM_API_KEY", "test-key")


class _FakeMsg:
    def __init__(self, role, text):
        self.role = role
        self._text = text
    def get_text_content(self):
        return self._text


class _FakeAgent:
    def __init__(self, msgs):
        self.full_trace = msgs


def test_summarize_trace_counts_roles_and_errors():
    from qea.worker_runtime import summarize_trace
    agent = _FakeAgent([
        _FakeMsg("assistant", "let me build it"),
        _FakeMsg("tool", "ok, file written"),
        _FakeMsg("assistant", "now verify"),
        _FakeMsg("tool", "Traceback (most recent call last): Error: boom"),
        _FakeMsg("user", "the task prompt"),
    ])
    mon = summarize_trace(agent)
    assert mon["turns"] == 2          # two assistant messages
    assert mon["tool_calls"] == 2     # two non-user tool/result messages
    assert mon["tool_errors"] == 1    # one carried an error marker


WEAK_DIR = Path(__file__).resolve().parent.parent / "qea" / "worker_gdpval_weak"
FULL_PROMPT = (Path(__file__).resolve().parent.parent
               / "qea" / "worker_gdpval" / "systemprompt.md").read_text()


def test_weak_seed_is_process_limited_not_capability_walled():
    # The weak seed must still load as a NexAU config and keep the shell tool
    # (so evolution CAN recover capability), but its prompt must be stripped of the
    # high-level finish-guidance / per-extension hints the full worker ships.
    from nexau import AgentConfig
    cfg = AgentConfig.from_yaml(config_path=WEAK_DIR / "agent.yaml")
    assert cfg is not None
    weak_prompt = (WEAK_DIR / "systemprompt.md").read_text()
    # headroom markers present in the FULL worker prompt, removed from the weak seed:
    assert "ls -la" not in weak_prompt              # finish/verify guidance removed
    assert "openpyxl" not in weak_prompt            # per-extension tool hints removed
    assert "Verify the file was written" not in weak_prompt
    assert len(weak_prompt) < len(FULL_PROMPT)      # strictly leaner
    # but the shell tool is still available (capability is recoverable by editing)
    tool_yaml = (WEAK_DIR / "tool_descriptions" / "run_shell_command.tool.yaml").read_text()
    assert "run_shell_command" in tool_yaml


def test_process_note_is_answer_free_and_file_signal_is_format_aware():
    from qea.debugger import process_note
    # a REQUIRED file is missing (format_ok False) -> flag "no deliverable file"
    n = process_note({"files": 0, "turns": 4, "tool_errors": 1, "format_ok": False})
    assert "no deliverable file" in n.lower()
    assert "4 turn" in n and "1 tool error" in n
    # TEXT-answer task (format_ok True, files=0 is NORMAL) -> do NOT flag a missing file
    txt = process_note({"files": 0, "turns": 40, "tool_errors": 2, "format_ok": True})
    assert "no deliverable file" not in txt.lower()
    assert "40 turn" in txt and "2 tool error" in txt
    # a healthy file-producing run
    ok = process_note({"files": 1, "turns": 11, "tool_errors": 0, "format_ok": True})
    assert "produced 1 file" in ok.lower() and "no deliverable file" not in ok.lower()
    # process notes carry only counts — never any answer/number-from-the-task content
    assert "$" not in n and "going-concern" not in n


def test_trace_fold_preserves_firewall():
    from qea.tasks import BTask
    from qea.verifier import TaskResult
    from qea.falsify import EvalSummary
    from qea.debugger import diagnose_b_pile

    class CriticLLM:
        def complete(self, prompt, *, role="judge"):
            # the OPEN-ENDED diagnosis call (Agent-Debugger `ask` style) asks for JSON
            # with a general_mechanism; it can name a TOOL to wire in.
            if "general_mechanism" in prompt:
                return ('{"root_cause": "worker cannot reach the filing", '
                        '"general_mechanism": "wire in the unbound retrieve_from_filing tool", '
                        '"kind": "tool"}')
            return "The deliverable omits the required reconciliation section."

    res = {"t1": TaskResult("t1", "Accountants and Auditors", "B", False, False, False, 0.3, 0.0,
                            None, criterion_verdicts={"1": False})}
    tasks = [BTask(task_id="t1", subtype="Accountants and Auditors", prompt="reconcile the ledger",
                   rubric="", rubric_items=[{"points": 1, "criterion": "reconciles to control total"}],
                   gold="SECRET-CONTROL-TOTAL-98765")]
    diag = diagnose_b_pile(EvalSummary(res, {"t1": "weak memo"}), tasks, llm=CriticLLM(),
                           traces={"t1": {"files": 0, "turns": 3, "tool_errors": 1}})
    payload = repr(diag.proposer_payload())
    assert "SECRET-CONTROL-TOTAL-98765" not in payload   # firewall holds with traces folded in
    assert "t1" in diag.predicted_fix_task_ids
    # open-ended diagnosis can point at a TOOL (what the retired 5-tag classifier couldn't)
    assert diag.mechanism_kind == "tool"
    assert "retrieve_from_filing" in diag.general_mechanism


def test_snapshot_and_diff_and_signature(tmp_path):
    from qea.evolve_runtime import snapshot_dir, dir_unified_diff, diff_signature, DirEdit
    src = tmp_path / "incumbent"
    (src / "tool_descriptions").mkdir(parents=True)
    (src / "systemprompt.md").write_text("original line\n")
    (src / "agent.yaml").write_text("name: w\n")

    snap = tmp_path / "snap"
    snapshot_dir(src, snap)
    assert (snap / "systemprompt.md").read_text() == "original line\n"

    # no edit yet -> empty diff, and a DirEdit over an empty diff has empty content
    assert dir_unified_diff(src, snap) == ""
    # edit the snapshot
    (snap / "systemprompt.md").write_text("original line\nADDED guidance\n")
    diff = dir_unified_diff(src, snap)
    assert "ADDED guidance" in diff and "systemprompt.md" in diff

    sig = diff_signature(diff)
    assert isinstance(sig, str) and len(sig) == 64           # sha256 hex
    # identical diff -> identical signature; different diff -> different
    assert diff_signature(diff) == sig
    assert diff_signature(diff + "x") != sig

    # DirEdit is the Edit-like shim the buffer + leakage guard consume
    de = DirEdit(diff)
    assert de.signature() == sig
    assert "ADDED guidance" in de.content            # leakage guard inspects .content
    assert de.summary                                 # non-empty human summary


def test_leakage_guard_blocks_dir_edit_that_pastes_answer():
    from qea.verifier import LeakageGuard
    from qea.evolve_runtime import DirEdit
    corpus = ["flags any going-concern triggers in the liquidity position"]
    guard = LeakageGuard(corpus, threshold=0.6)
    leaked_diff = ("--- a/systemprompt.md\n+++ b/systemprompt.md\n"
                   "+Always flags any going-concern triggers in the liquidity position.\n")
    assert guard.is_leak(DirEdit(leaked_diff)) is True
    ok_diff = ("--- a/systemprompt.md\n+++ b/systemprompt.md\n"
               "+After producing the file, list the directory to verify it was written.\n")
    assert guard.is_leak(DirEdit(ok_diff)) is False


def test_evolve_agent_config_loads():
    from nexau import AgentConfig
    from qea.evolve_runtime import EVOLVE_DIR
    cfg = AgentConfig.from_yaml(config_path=EVOLVE_DIR / "agent.yaml")
    assert cfg is not None


class _FakeEval:
    """Benchmark-agnostic fake Evaluator: score depends on whether the worker prompt
    was improved (the evolve agent appends a marker the fake worker echoes)."""
    def __init__(self, base=0.50, improved=0.90):
        self.base, self.improved = base, improved
    def evaluate(self, task, worker_run, out_dir=None):
        from qea.evaluator import TaskEval
        s = self.improved if "improved=True" in worker_run.deliverable_text else self.base
        return TaskEval(s, s, True, worker_run.deliverable_text, {"1": s > 0.6}, 0.0)


def test_levelb_loop_keeps_improving_edit_offline(tmp_path, monkeypatch):
    import qea.loop_levelb as L
    import qea.worker_runtime
    from qea.worker_runtime import WorkerRun
    from qea.tasks import BTask

    tasks = [BTask(task_id="t1", subtype="Accountants and Auditors", prompt="p", rubric="",
                   rubric_items=[{"points": 1, "criterion": "c"}], gold="g")]

    class StubLLM:  # the firewalled debugger's critic + classify calls
        def complete(self, prompt, *, role="judge", **kw):
            return ('{"root_cause_tag":"WrongStructure","target_slot":"prompt"}'
                    if "Classify" in prompt else "omits the required structure")

    # worker: returns a fixed deliverable; "improved=True" once the prompt is edited.
    def fake_run_worker(task, worker_dir, run_dir):
        improved = "IMPROVED" in (worker_dir / "systemprompt.md").read_text()
        return WorkerRun(f"deliverable improved={improved}", [], {"files": 1, "turns": 5, "tool_errors": 0})
    monkeypatch.setattr(qea.worker_runtime, "run_worker", fake_run_worker)

    # evolve agent: appends the IMPROVED marker + predicts it fixes t1
    def fake_run_evolve(snapshot_dir_path, diag, run_dir, *, edit_history="", evidence_dir=None):
        sp = snapshot_dir_path / "systemprompt.md"
        sp.write_text(sp.read_text() + "\nIMPROVED: verify the file before finishing.\n")
        return {"final_text": "added verify guidance", "trace": {"turns": 2},
                "prediction": {"predicted_fixes": ["t1"], "risk_tasks": []}}
    monkeypatch.setattr(L, "run_evolve_agent", fake_run_evolve)

    seed = tmp_path / "seed"; (seed / "tool_descriptions").mkdir(parents=True)
    (seed / "agent.yaml").write_text("name: w\n")
    (seed / "systemprompt.md").write_text("do the task\n")
    cfg = L.LevelBConfig(n_iters=1, k=1, n_tasks=1, results_dir=str(tmp_path / "res"),
                         seed_worker_dir=str(seed))

    res = L.run_levelb(cfg, _tasks=tasks, _evaluator=_FakeEval(), _llm=StubLLM())
    assert res.n_kept == 1                                   # the +0.40 edit beats the noise floor
    assert res.records[0].verdict == "EFFECTIVE"            # predicted t1 fixed, t1 flipped
    assert res.records[0].improved == ["t1"]
    assert res.mean_score_trajectory[-1] > res.mean_score_trajectory[0]
    assert "IMPROVED" in (Path(res.final_worker_dir) / "systemprompt.md").read_text()


def test_levelb_loop_rolls_back_non_improving_edit(tmp_path, monkeypatch):
    import qea.loop_levelb as L
    import qea.worker_runtime
    from qea.worker_runtime import WorkerRun
    from qea.tasks import BTask

    tasks = [BTask(task_id="t1", subtype="x", prompt="p", rubric="",
                   rubric_items=[{"points": 1, "criterion": "c"}], gold="g")]

    class StubLLM:
        def complete(self, prompt, *, role="judge", **kw):
            return ('{"root_cause_tag":"WrongStructure","target_slot":"prompt"}'
                    if "Classify" in prompt else "omits the required structure")

    monkeypatch.setattr(qea.worker_runtime, "run_worker",
                        lambda task, wd, rd: WorkerRun("same", [], {"files": 1, "turns": 5, "tool_errors": 0}))

    # evolve makes a real (but useless) change so the diff is non-empty; predicts t1
    def ev(snap, diag, rd, *, edit_history="", evidence_dir=None):
        sp = snap / "systemprompt.md"; sp.write_text(sp.read_text() + "\nnoise edit\n")
        return {"final_text": "x", "trace": {}, "prediction": {"predicted_fixes": ["t1"], "risk_tasks": []}}
    monkeypatch.setattr(L, "run_evolve_agent", ev)

    seed = tmp_path / "seed"; (seed / "tool_descriptions").mkdir(parents=True)
    (seed / "agent.yaml").write_text("name: w\n"); (seed / "systemprompt.md").write_text("do it\n")
    cfg = L.LevelBConfig(n_iters=1, k=1, n_tasks=1, results_dir=str(tmp_path / "res"),
                         seed_worker_dir=str(seed))
    # a flat evaluator: score never moves -> nothing flips -> INEFFECTIVE -> rolled back
    res = L.run_levelb(cfg, _tasks=tasks, _evaluator=_FakeEval(base=0.50, improved=0.50),
                       _llm=StubLLM())
    assert res.n_kept == 0 and res.n_rolled_back == 1        # flat score -> rolled back
    assert res.records[0].verdict == "INEFFECTIVE"


def test_format_gate_keyed_to_gold_extension():
    from qea.grading.format_gate import format_ok, apply_gate
    from qea.tasks import BTask

    # text-gold task (no required ext) -> always ok, even with no files
    txt = BTask(task_id="t", subtype="x", prompt="p", rubric="", deliverable_exts=[])
    assert format_ok(txt, []) is True
    assert apply_gate(0.8, txt, []) == (0.8, True)

    # xlsx-gold task: must produce an .xlsx
    xls = BTask(task_id="t", subtype="x", prompt="p", rubric="", deliverable_exts=[".xlsx"])
    assert format_ok(xls, ["/tmp/report.xlsx"]) is True
    assert apply_gate(0.83, xls, ["/tmp/report.xlsx"]) == (0.83, True)
    # wrong container (asked xlsx, produced pdf) -> gated to 0
    assert format_ok(xls, ["/tmp/report.pdf"]) is False
    assert apply_gate(0.83, xls, ["/tmp/report.pdf"]) == (0.0, False)
    # no file at all (text answer) -> gated to 0
    assert apply_gate(0.83, xls, []) == (0.0, False)

    # multi-ext gold: any one matching extension satisfies it
    both = BTask(task_id="t", subtype="x", prompt="p", rubric="", deliverable_exts=[".pdf", ".xlsx"])
    assert format_ok(both, ["/tmp/a.xlsx"]) is True


def test_gdpval_loader_populates_deliverable_exts():
    # The GDPval loader must attach the gold deliverable extension(s) to each task.
    from qea.tasks import _gold_deliverable_exts
    assert _gold_deliverable_exts(["deliverable_files/abc/Fall Music Tour Output.xlsx"]) == [".xlsx"]
    assert _gold_deliverable_exts(None) == []          # text deliverable -> no required ext
    assert _gold_deliverable_exts(["a/x.pdf", "b/y.pptx"]) == [".pdf", ".pptx"]


def test_rubric_text_evaluator_scores_and_no_op_gate():
    from qea.evaluator import RubricTextEvaluator
    from qea.worker_runtime import WorkerRun
    from qea.tasks import BTask

    class JudgeLLM:
        def complete(self, prompt, *, role="judge", **kw):
            return "1: yes\n2: no"            # 1 of 2 criteria satisfied -> 0.5
    task = BTask(task_id="f0", subtype="x", prompt="p", rubric="",
                 rubric_items=[{"points": 1, "criterion": "a"}, {"points": 1, "criterion": "b"}])
    te = RubricTextEvaluator(JudgeLLM(), k=1).evaluate(task, WorkerRun("my answer", [], {}), None)
    assert 0.0 <= te.content_score <= 1.0
    assert te.format_ok is True and te.gated_score == te.content_score   # text-gold -> gate no-op
    assert te.deliverable_text == "my answer"


def test_multimodal_evaluator_applies_format_gate():
    from qea.evaluator import MultimodalEvaluator
    from qea.worker_runtime import WorkerRun
    from qea.grading.multimodal_judge import GradeResult
    from qea.tasks import BTask
    from types import SimpleNamespace

    ev = MultimodalEvaluator(llm=None)
    ev.judge = SimpleNamespace(grade=lambda task, rendered: GradeResult(task.task_id, 0.8, 0.8, {"1": True}, 0.0, False))
    # stub render so no LibreOffice/PyMuPDF is needed (evaluate imports it at call time)
    monkey = SimpleNamespace(text="t", extracted_text="t", images=[], degraded=[])
    import qea.grading.render as R
    orig = R.render
    R.render = lambda text, files, out_dir: monkey
    try:
        xls = BTask(task_id="g0", subtype="x", prompt="p", rubric="", deliverable_exts=[".xlsx"])
        ok = ev.evaluate(xls, WorkerRun("t", ["/tmp/report.xlsx"], {}), "/tmp/out")
        assert ok.gated_score == 0.8 and ok.format_ok is True
        miss = ev.evaluate(xls, WorkerRun("t", ["/tmp/report.pdf"], {}), "/tmp/out")
        assert miss.content_score == 0.8 and miss.gated_score == 0.0 and miss.format_ok is False
    finally:
        R.render = orig


def test_prediction_parsing_handles_present_absent_malformed():
    from qea.evolve_runtime import _parse_prediction
    good = _parse_prediction('done.\n{"predicted_fixes": ["t1", "t2"], "risk_tasks": ["t3"]}')
    assert good == {"predicted_fixes": ["t1", "t2"], "risk_tasks": ["t3"]}
    absent = _parse_prediction("I changed the prompt, no JSON here.")
    assert absent == {"predicted_fixes": [], "risk_tasks": []}
    malformed = _parse_prediction('{"predicted_fixes": [oops')
    assert malformed == {"predicted_fixes": [], "risk_tasks": []}


def test_classify_verdict_table():
    from qea.loop_levelb import _classify
    from qea.evolve_runtime import DirEdit
    from qea.evaluator import TaskEval

    def evals(scores):  # tid -> gated score
        return {t: TaskEval(s, s, True, "", {}, 0.0) for t, s in scores.items()}

    inc = evals({"a": 0.2, "b": 0.5, "c": 0.5})
    # EFFECTIVE: predicted a fixed, a flips, nothing regresses
    e = DirEdit("d", predicted_fixes=["a"], risk_tasks=[])
    assert _classify(e, inc, evals({"a": 0.9, "b": 0.5, "c": 0.5}), 0.05)["verdict"] == "EFFECTIVE"
    # HARMFUL: nothing predicted-fixed flipped, an unattributed task regressed
    e2 = DirEdit("d", predicted_fixes=["a"], risk_tasks=[])
    assert _classify(e2, inc, evals({"a": 0.2, "b": 0.1, "c": 0.5}), 0.05)["verdict"] == "HARMFUL"
    # MIXED: a predicted fix flips AND an unattributed task regresses
    e3 = DirEdit("d", predicted_fixes=["a"], risk_tasks=[])
    assert _classify(e3, inc, evals({"a": 0.9, "b": 0.1, "c": 0.5}), 0.05)["verdict"] == "MIXED"
    # INEFFECTIVE: nothing moves
    e4 = DirEdit("d", predicted_fixes=["a"], risk_tasks=[])
    assert _classify(e4, inc, evals({"a": 0.2, "b": 0.5, "c": 0.5}), 0.05)["verdict"] == "INEFFECTIVE"


def test_nexau_reference_guide_is_loadable_and_substrate_only():
    from qea.evolve_runtime import _read_reference
    ref = _read_reference()
    assert ref and "binding:" in ref and "tools:" in ref      # (b) reference present
    assert "max_iterations" in ref and "tool_descriptions" in ref


def test_build_evidence_corpus_includes_failure_not_gold(tmp_path):
    from qea.loop_levelb import _build_evidence
    from qea.evaluator import TaskEval
    from qea.tasks import BTask

    tasks = [BTask(task_id="t1", subtype="x", prompt="p", rubric="",
                   rubric_items=[{"points": 1, "criterion": "reconciles to the control total"},
                                 {"points": 1, "criterion": "lists assumptions"}],
                   gold="SECRET-GOLD-ANSWER-42")]
    evals = {"t1": TaskEval(0.2, 0.2, False, "my weak memo", {"1": False, "2": True}, 0.0)}
    traces = {"t1": {"files": 0, "turns": 3, "tool_errors": 1, "trace_path": "/x/trace.txt"}}
    diag = {"root_cause_tag": "WrongStructure", "overview": "structure missing"}

    ed = _build_evidence(tmp_path / "ev", diag, evals, traces, {"t1": "my weak memo"}, tasks, [])
    overview = (ed / "overview.md").read_text()
    assert "reconciles to the control total" in overview      # the FAILED criterion (verdict False)
    assert "lists assumptions" not in overview                # the passed criterion is not listed
    assert "my weak memo" in overview                         # worker's own deliverable
    assert "SECRET-GOLD-ANSWER-42" not in overview            # gold is NEVER written
    assert (ed / "evolution_history.md").exists()


def test_levelb_ahe_corpus_mode_passes_evidence_dir(tmp_path, monkeypatch):
    import qea.loop_levelb as L
    import qea.worker_runtime
    from qea.worker_runtime import WorkerRun
    from qea.tasks import BTask

    tasks = [BTask(task_id="t1", subtype="x", prompt="p", rubric="",
                   rubric_items=[{"points": 1, "criterion": "c"}], gold="g")]
    monkeypatch.setattr(qea.worker_runtime, "run_worker",
                        lambda task, wd, rd: WorkerRun("weak ans", [], {"files": 0, "turns": 2, "tool_errors": 0}))

    captured = {}
    def ev(snap, diag, rd, *, edit_history="", evidence_dir=None):
        captured["evidence_dir"] = evidence_dir
        return {"final_text": "x", "trace": {}, "prediction": {"predicted_fixes": [], "risk_tasks": []}}
    monkeypatch.setattr(L, "run_evolve_agent", ev)

    class StubLLM:
        def complete(self, prompt, *, role="judge", **kw):
            return ('{"root_cause_tag":"WrongStructure","target_slot":"prompt"}'
                    if "Classify" in prompt else "omits structure")

    seed = tmp_path / "seed"; (seed / "tool_descriptions").mkdir(parents=True)
    (seed / "agent.yaml").write_text("name: w\n"); (seed / "systemprompt.md").write_text("do it\n")
    cfg = L.LevelBConfig(n_iters=1, k=1, n_tasks=1, results_dir=str(tmp_path / "res"),
                         seed_worker_dir=str(seed), evidence_mode="ahe_corpus")
    # failing task (0.3 < 0.6) so it lands in the evidence overview
    L.run_levelb(cfg, _tasks=tasks, _evaluator=_FakeEval(base=0.3, improved=0.3), _llm=StubLLM())
    assert captured["evidence_dir"] is not None
    assert (Path(captured["evidence_dir"]) / "overview.md").exists()


def test_make_benchmark_routes_to_the_right_evaluator():
    from qea.benchmark import make_benchmark
    from qea.evaluator import RubricTextEvaluator
    fab = make_benchmark("fab", llm=None, k=2)
    assert fab.name == "fab_v2" and isinstance(fab.evaluator, RubricTextEvaluator)
    assert fab.answer_corpus and all(isinstance(c, str) for c in fab.answer_corpus)
    with pytest.raises(ValueError):
        make_benchmark("nonsense")


def test_evaluate_dir_tolerates_worker_failure(tmp_path, monkeypatch):
    # A crashing worker (e.g. NexAU empty-response after retries) must not kill the
    # run: the task scores 0, the error is recorded in the trace, the loop continues.
    import qea.loop_levelb as L
    import qea.worker_runtime
    from qea.tasks import BTask

    tasks = [BTask(task_id="t1", subtype="x", prompt="p", rubric="",
                   rubric_items=[{"points": 1, "criterion": "c"}]),
             BTask(task_id="t2", subtype="x", prompt="p", rubric="",
                   rubric_items=[{"points": 1, "criterion": "c"}])]

    def boom(task, wd, rd):
        if task.task_id == "t1":
            raise RuntimeError("Error in agent execution: No response content or tool calls")
        from qea.worker_runtime import WorkerRun
        return WorkerRun("ok answer", [], {"files": 0, "turns": 3, "tool_errors": 0})
    monkeypatch.setattr(qea.worker_runtime, "run_worker", boom)

    evals, traces, deliverables, mean = L.evaluate_dir(
        tmp_path / "w", tasks, _FakeEval(base=0.7, improved=0.7), tmp_path / "r")
    assert evals["t1"].gated_score == 0.0 and evals["t1"].format_ok is False
    assert "RuntimeError" in traces["t1"]["error"]
    assert evals["t2"].gated_score == 0.7                 # the healthy task still scored
    assert mean == 0.35                                    # (0.0 + 0.7) / 2


@pytest.mark.skipif(os.environ.get("QEA_LEVELB_SMOKE") != "1",
                    reason="set QEA_LEVELB_SMOKE=1 to run the real-API NexAU Level-B smoke test")
def test_levelb_smoke_one_task_one_iter(tmp_path):
    # End-to-end on ONE real FAB task, ONE iteration, against the real NexAU weak
    # worker + evolve agent. Proves the wiring runs; makes no headroom claim.
    import run as runmod
    runmod._load_dotenv()
    from qea.loop_levelb import LevelBConfig, run_levelb
    cfg = LevelBConfig(n_iters=1, k=1, n_tasks=1, benchmark="fab",
                       seed_worker_dir="qea/worker_fab_weak", results_dir=str(tmp_path / "res"))
    res = run_levelb(cfg)
    assert res.n_tasks == 1
    assert len(res.mean_score_trajectory) >= 1
    assert Path(res.final_worker_dir).exists()
    assert (Path(cfg.results_dir) / "iter_001" / "manifest.json").exists()


def test_build_evidence_attempt_archive_and_score_matrix(tmp_path):
    """The anti-fixation archive: past attempt diffs land in past_edits/ labeled with
    outcome + helped/hurt, the task x attempt score matrix is written, and the
    directive tells the agent to merge components instead of retrying falsified
    approaches."""
    from qea.loop_levelb import _build_evidence
    from qea.evaluator import TaskEval
    from qea.tasks import BTask

    tasks = [BTask(task_id="t1", subtype="x", prompt="p", rubric="",
                   rubric_items=[{"points": 1, "criterion": "c"}], gold="g")]
    evals = {"t1": TaskEval(0.2, 0.2, False, "d", {"1": False}, 0.0)}
    traces = {"t1": {"files": 0, "turns": 1, "tool_errors": 0}}
    past = [{"name": "iter_001", "kept": False, "verdict": "MIXED",
             "summary": "edit tools/excel.py (+40/-0)", "improved": ["taskA"],
             "regressed": ["taskB"], "diff": "+++ b/tools/excel.py\n+def read_xlsx():"}]
    scores = {"seed": {"t1": 0.2, "t2": 0.9}, "iter1": {"t1": 0.5, "t2": 0.4}}

    ed = _build_evidence(tmp_path / "ev", {"root_cause_tag": "x", "overview": "o"}, evals,
                         traces, {"t1": "d"}, tasks, [], attempt_scores=scores,
                         past_edits=past)
    pe = (ed / "past_edits" / "iter_001.diff").read_text()
    assert "MIXED (rolled back)" in pe and "helped: taskA" in pe and "hurt: taskB" in pe
    assert "def read_xlsx" in pe                       # the actual diff body
    matrix = (ed / "archive_scores.md").read_text()
    assert "| t1 | 0.20 | 0.50 |" in matrix            # per-task columns in order
    hist = (ed / "evolution_history.md").read_text()
    assert "Approach constraints" not in hist          # no attempts yet -> no directive

    ed2 = _build_evidence(tmp_path / "ev2", {"root_cause_tag": "x", "overview": "o"}, evals,
                          traces, {"t1": "d"}, tasks, [], prior_history="- prior run iter_001: e -> MIXED",
                          attempt_scores=scores, past_edits=past)
    hist2 = (ed2 / "evolution_history.md").read_text()
    assert "Approach constraints" in hist2 and "MERGE" in hist2


def test_edit_history_and_prior_manifests_carry_helped_hurt(tmp_path):
    """History lines (in-run and cross-leg) expose WHICH tasks each falsified edit
    helped/hurt — the opaque '-> MIXED' label alone did not stop re-proposals."""
    import json as _json
    from qea.loop_levelb import LevelBRecord, _edit_history, _load_prior_history, _load_prior_edits

    r = LevelBRecord(1, False, "MIXED", False, "edit tools/excel.py", "tag", 0.5, 0.52,
                     improved=["aaaa1111-x"], regressed=["bbbb2222-y"])
    h = _edit_history([r])
    assert "helped: aaaa1111" in h and "hurt: bbbb2222" in h and "-> MIXED" in h

    it = tmp_path / "iter_001"; it.mkdir()
    (it / "manifest.json").write_text(_json.dumps({
        "kept": False, "verdict": "HARMFUL", "edit_summary": "edit x",
        "improved": [], "regressed": ["cccc3333"]}))
    (it / "edit.diff").write_text("+++ b/x\n+z")
    ph = _load_prior_history(tmp_path)
    assert "hurt: cccc3333" in ph
    edits = _load_prior_edits(tmp_path)
    assert edits[0]["name"].startswith("prior_") and edits[0]["name"].endswith("_iter_001")
    assert edits[0]["diff"].startswith("+++")

    # multi-leg spec: comma-separated dirs are concatenated oldest-first
    leg2 = tmp_path / "leg2" / "iter_001"; leg2.mkdir(parents=True)
    (leg2 / "manifest.json").write_text(_json.dumps({
        "kept": True, "verdict": "EXPECTED", "edit_summary": "edit y"}))
    (leg2 / "edit.diff").write_text("+++ b/y\n+w")
    both = _load_prior_edits(f"{tmp_path},{tmp_path / 'leg2'}")
    assert len(both) == 2 and both[1]["kept"]
    hist = _load_prior_history(f"{tmp_path},{tmp_path / 'leg2'}")
    assert "leg2/iter_001" in hist and "-> KEPT" in hist


def test_confirm_band_second_eval_gates_near_floor_keeps(tmp_path, monkeypatch):
    """decision-(2): a near-floor keep must survive a second independent candidate
    eval (measured GDPval delta sigma ~0.054 > the 0.05 floor). Flaky gain -> the
    confirmation eval disagrees and the edit is rolled back; stable gain -> kept."""
    import qea.loop_levelb as L
    import qea.worker_runtime
    from qea.worker_runtime import WorkerRun
    from qea.tasks import BTask

    tasks = [BTask(task_id=f"t{i}", subtype="x", prompt="p", rubric="",
                   rubric_items=[{"points": 1, "criterion": "c"}], gold="g") for i in range(2)]
    monkeypatch.setattr(qea.worker_runtime, "run_worker",
                        lambda task, wd, rd: WorkerRun(
                            "improved=True" if "MARKER" in (Path(wd) / "systemprompt.md").read_text()
                            else "base", [], {"files": 0, "turns": 1, "tool_errors": 0}))

    def ev(snap, diag, rd, *, edit_history="", evidence_dir=None):
        sp = Path(snap) / "systemprompt.md"
        sp.write_text(sp.read_text() + "\nMARKER\n")
        return {"final_text": "e", "trace": {},
                "prediction": {"predicted_fixes": [t.task_id for t in tasks], "risk_tasks": []}}
    monkeypatch.setattr(L, "run_evolve_agent", ev)

    class StubLLM:
        def complete(self, prompt, *, role="judge", **kw):
            return ('{"root_cause_tag":"WrongStructure","target_slot":"prompt"}'
                    if "Classify" in prompt else "weak")

    class FlakyEval:
        """Improved worker scores 0.58 on the first candidate eval (clears the 0.05
        floor over base 0.50) but 0.50 on the confirmation -> average 0.54 misses."""
        def __init__(self):
            self.improved_evals = 0
        def evaluate(self, task, worker_run, out_dir=None):
            from qea.evaluator import TaskEval
            if "improved=True" in worker_run.deliverable_text:
                self.improved_evals += 1
                s = 0.58 if self.improved_evals <= 2 else 0.50  # 2 tasks/eval
            else:
                s = 0.50
            return TaskEval(s, s, True, worker_run.deliverable_text, {"1": s > 0.6}, 0.0)

    def mkseed(d):
        seed = d / "seed"; (seed / "tool_descriptions").mkdir(parents=True)
        (seed / "agent.yaml").write_text("name: w\n"); (seed / "systemprompt.md").write_text("do\n")
        return seed

    cfg = L.LevelBConfig(n_iters=1, k=1, n_tasks=2, results_dir=str(tmp_path / "res1"),
                         seed_worker_dir=str(mkseed(tmp_path / "a")), noise_margin=0.05,
                         confirm_band=0.10)
    res = L.run_levelb(cfg, _tasks=tasks, _evaluator=FlakyEval(), _llm=StubLLM())
    assert res.n_kept == 0 and res.n_rolled_back == 1   # confirmation falsified the gain

    class StableEval(FlakyEval):
        def evaluate(self, task, worker_run, out_dir=None):
            from qea.evaluator import TaskEval
            s = 0.70 if "improved=True" in worker_run.deliverable_text else 0.50
            return TaskEval(s, s, True, worker_run.deliverable_text, {"1": s > 0.6}, 0.0)

    cfg2 = L.LevelBConfig(n_iters=1, k=1, n_tasks=2, results_dir=str(tmp_path / "res2"),
                          seed_worker_dir=str(mkseed(tmp_path / "b")), noise_margin=0.05,
                          confirm_band=0.10)
    res2 = L.run_levelb(cfg2, _tasks=tasks, _evaluator=StableEval(), _llm=StubLLM())
    assert res2.n_kept == 1                              # 0.70 twice -> confirmed


def test_evaluate_dir_n_samples_averages_and_caches_per_sample(tmp_path, monkeypatch):
    """n_samples > 1: per-task score is the mean over independent worker runs
    (variance reduction for catastrophically-flippy weak workers — measured per-task
    repeat sd up to ~0.29 on GDPval), samples cache under distinct keys, and the
    best sample supplies the deliverable."""
    import qea.loop_levelb as L
    import qea.worker_runtime
    from qea.worker_runtime import WorkerRun
    from qea.tasks import BTask

    tasks = [BTask(task_id="t1", subtype="x", prompt="p", rubric="",
                   rubric_items=[{"points": 1, "criterion": "c"}], gold="g")]
    calls = {"n": 0}

    def flippy(task, wd, rd):
        calls["n"] += 1
        return WorkerRun(f"answer-{calls['n']}", [], {"files": 0, "turns": 1, "tool_errors": 0})
    monkeypatch.setattr(qea.worker_runtime, "run_worker", flippy)

    class FlipEval:
        """Scores alternate 0.9 / 0.1 by sample — a catastrophic-flip task."""
        def evaluate(self, task, worker_run, out_dir=None):
            from qea.evaluator import TaskEval
            s = 0.9 if int(worker_run.deliverable_text.split("-")[1]) % 2 else 0.1
            return TaskEval(s, s, True, worker_run.deliverable_text, {"1": True}, 0.0)

    wd = tmp_path / "w"; (wd / "tool_descriptions").mkdir(parents=True)
    (wd / "agent.yaml").write_text("name: w\n"); (wd / "systemprompt.md").write_text("do\n")
    cache = tmp_path / "cache"
    evals, traces, deliv, mean = L.evaluate_dir(
        wd, tasks, FlipEval(), tmp_path / "run", cache_dir=cache, n_samples=2)
    assert abs(evals["t1"].gated_score - 0.5) < 1e-9      # mean of 0.9 and 0.1
    assert abs(evals["t1"].variance - 0.4) < 1e-9         # pstdev of [0.9, 0.1]
    assert deliv["t1"] == "answer-1"                      # best sample (0.9) wins
    names = sorted(f.name for f in cache.iterdir())
    assert len(names) == 2 and names[1].endswith("__s1.json")  # s0 keeps legacy name
    assert "__s" not in names[0]

    # resume: a second call re-runs nothing (both samples cached)
    calls["n"] = 0
    evals2, _, _, _ = L.evaluate_dir(wd, tasks, FlipEval(), tmp_path / "run2",
                                     cache_dir=cache, n_samples=2)
    assert calls["n"] == 0 and abs(evals2["t1"].gated_score - 0.5) < 1e-9


def test_decide_keep_paired_gate():
    """Protocol-v2 keep gate: paired bootstrap + stability penalty."""
    from qea.falsify import decide_keep_paired

    # uniform genuine gain on every task -> kept (CI above 0, mean above floor)
    inc = {f"t{i}": 0.5 for i in range(8)}
    cand = {f"t{i}": 0.6 for i in range(8)}
    d = decide_keep_paired(inc, cand, noise_margin=0.05)
    assert d["kept"] and d["mean_delta"] == 0.1 and d["ci_lo"] > 0

    # one big win, rest tiny losses: mean clears the floor but the direction does
    # not survive resampling -> rolled back (this is the noise-driven keep the
    # soft gate would have accepted)
    inc2 = {f"t{i}": 0.5 for i in range(8)}
    cand2 = {**{f"t{i}": 0.49 for i in range(8)}, "t0": 1.0}
    from qea.falsify import decide_keep_soft
    m2 = sum(cand2.values()) / 8
    assert decide_keep_soft(0.5, m2, 0.05)          # legacy gate would keep
    d2 = decide_keep_paired(inc2, cand2, noise_margin=0.05)
    assert not d2["kept"] and d2["ci_lo"] <= 0      # paired gate rejects

    # stability penalty: same means, candidate much noisier across samples -> rejected
    vars_inc = {f"t{i}": 0.0 for i in range(8)}
    vars_cand = {f"t{i}": 0.4 for i in range(8)}
    d3 = decide_keep_paired(inc, cand, noise_margin=0.05, stability_lambda=0.5,
                            inc_vars=vars_inc, cand_vars=vars_cand)
    assert not d3["kept"] and d3["objective_delta"] < 0.05

    # no common tasks -> safe reject
    assert not decide_keep_paired({"a": 1.0}, {"b": 1.0})["kept"]


def test_confirm_tasks_heldout_gate_rejects_overfit_keep(tmp_path, monkeypatch):
    """Regimes-style promote-time gate: an edit that wins on the optimize pool but
    regresses on the held-out confirm batch is rolled back as overfit."""
    import qea.loop_levelb as L
    import qea.worker_runtime
    from qea.worker_runtime import WorkerRun
    from qea.tasks import BTask

    # 2 optimize tasks + 1 confirm task
    tasks = [BTask(task_id=f"t{i}", subtype="x", prompt="p", rubric="",
                   rubric_items=[{"points": 1, "criterion": "c"}], gold="g") for i in range(3)]
    monkeypatch.setattr(qea.worker_runtime, "run_worker",
                        lambda task, wd, rd: WorkerRun(
                            f"{task.task_id}|improved={'MARKER' in (Path(wd) / 'systemprompt.md').read_text()}",
                            [], {"files": 0, "turns": 1, "tool_errors": 0}))

    def ev(snap, diag, rd, *, edit_history="", evidence_dir=None):
        sp = Path(snap) / "systemprompt.md"
        sp.write_text(sp.read_text() + "\nMARKER\n")
        return {"final_text": "e", "trace": {},
                "prediction": {"predicted_fixes": ["t0", "t1"], "risk_tasks": []}}
    monkeypatch.setattr(L, "run_evolve_agent", ev)

    class StubLLM:
        def complete(self, prompt, *, role="judge", **kw):
            return ('{"root_cause_tag":"WrongStructure","target_slot":"prompt"}'
                    if "Classify" in prompt else "weak")

    class OverfitEval:
        """Improved worker: big win on t0/t1 (optimize) but tanks on t2 (confirm)."""
        def evaluate(self, task, worker_run, out_dir=None):
            from qea.evaluator import TaskEval
            tid = worker_run.deliverable_text.split("|")[0]
            improved = "improved=True" in worker_run.deliverable_text
            if improved:
                s = 0.9 if tid in ("t0", "t1") else 0.1
            else:
                s = 0.5
            return TaskEval(s, s, True, worker_run.deliverable_text, {"1": s > 0.6}, 0.0)

    seed = tmp_path / "seed"; (seed / "tool_descriptions").mkdir(parents=True)
    (seed / "agent.yaml").write_text("name: w\n"); (seed / "systemprompt.md").write_text("do\n")
    cfg = L.LevelBConfig(n_iters=1, k=1, n_tasks=2, results_dir=str(tmp_path / "res"),
                         seed_worker_dir=str(seed), noise_margin=0.05, confirm_tasks=1)
    res = L.run_levelb(cfg, _tasks=tasks, _evaluator=OverfitEval(), _llm=StubLLM())
    assert res.n_kept == 0 and res.n_rolled_back == 1   # confirm gate caught the overfit


def test_judge_reason_mode_stays_out_of_verdicts(tmp_path):
    """Reason-mode judge responses: verdicts normalize to plain bools (debugger
    consumers test `is False`), reasons go to judge_reasons.json only."""
    from qea.verifier import build_rubric_prompt, rubric_reasons, score_rubric
    from qea.tasks import BTask

    task = BTask(task_id="t", subtype="x", prompt="p", rubric="",
                 rubric_items=[{"points": 2, "criterion": "sums to control total"},
                               {"points": 1, "criterion": "lists assumptions"}], gold="g")
    p = build_rubric_prompt(task, "d", task.rubric_items, with_reasons=True)
    assert '"reason"' in p and "ONE short sentence" in p
    p0 = build_rubric_prompt(task, "d", task.rubric_items)
    assert '"reason"' not in p0                       # legacy shape unchanged

    raw = ('{"1": {"pass": true, "reason": "total matches"}, '
           '"2": {"pass": false, "reason": "no assumptions section"}}')
    frac, verdicts = score_rubric(raw, task.rubric_items)
    assert verdicts == {"1": True, "2": False}        # bools only, never dicts
    assert abs(frac - 2 / 3) < 1e-9
    rr = rubric_reasons(raw, task.rubric_items)
    assert rr["2"]["reason"] == "no assumptions section"
    assert rr["2"]["criterion"] == "lists assumptions"

    # evaluator writes the debug file next to the eval outputs
    from qea.evaluator import _write_judge_reasons
    _write_judge_reasons(tmp_path / "task1", rr)
    import json as _json
    saved = _json.loads((tmp_path / "task1" / "judge_reasons.json").read_text())
    assert saved["1"]["pass"] is True

    # firewall check: the evidence builder never touches judge_reasons content
    from qea.loop_levelb import _build_evidence
    from qea.evaluator import TaskEval
    evals = {"t": TaskEval(0.2, 0.2, False, "d", verdicts, 0.0)}
    ed = _build_evidence(tmp_path / "ev", {"root_cause_tag": "x", "overview": "o"},
                         evals, {"t": {}}, {"t": "d"}, [task], [])
    assert "no assumptions section" not in (ed / "overview.md").read_text()


def test_ssb_loader_and_evaluator_offline(tmp_path, monkeypatch):
    """SpreadsheetBench integration: per-case tasks, answer file never in the task
    surface, official checker drives a binary TaskEval."""
    import json
    import openpyxl
    import qea.bench_ssb as B

    # fabricate a mini split: 1 task, 2 test cases
    base = tmp_path / "raw" / "all_data_912_v0.1"
    sp = base / "spreadsheet" / "77"
    sp.mkdir(parents=True)
    for k in (1, 2):
        wb = openpyxl.Workbook(); ws = wb.active
        ws["A1"] = 10 * k; wb.save(sp / f"{k}_77_input.xlsx")
        wb2 = openpyxl.Workbook(); ws2 = wb2.active
        ws2["A1"] = 10 * k; ws2["B1"] = 10 * k + 1   # answer: B1 = A1 + 1
        wb2.save(sp / f"{k}_77_answer.xlsx")
    (base / "dataset.json").write_text(json.dumps([{
        "id": "77", "instruction": "Put A1+1 into B1.",
        "instruction_type": "Cell-Level Manipulation", "answer_position": "B1"}]))
    monkeypatch.setattr(B, "_ROOT", tmp_path)

    tasks = B.load_ssb(split="912")
    assert [t.task_id for t in tasks] == ["77_tc1", "77_tc2"]
    t1 = tasks[0]
    assert t1.reference_files[0].endswith("1_77_input.xlsx")
    assert "answer.xlsx" not in t1.prompt and "answer" not in str(t1.reference_files)  # leakage
    assert t1.deliverable_exts == [".xlsx"] and "1_77_output.xlsx" in t1.prompt

    # evaluator: correct output -> 1.0; wrong -> 0.0; missing -> 0.0 + format fail
    from qea.worker_runtime import WorkerRun
    ev = B.SSBEvaluator(recalc=False)
    good = tmp_path / "1_77_output.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws["A1"] = 10; ws["B1"] = 11; wb.save(good)
    r = ev.evaluate(t1, WorkerRun("done", [good], {}))
    assert r.gated_score == 1.0 and r.format_ok

    bad = tmp_path / "bad" / "1_77_output.xlsx"; bad.parent.mkdir()
    wb = openpyxl.Workbook(); ws = wb.active; ws["A1"] = 10; ws["B1"] = 99; wb.save(bad)
    assert ev.evaluate(t1, WorkerRun("done", [bad], {})).gated_score == 0.0
    miss = ev.evaluate(t1, WorkerRun("done", [], {}))
    assert miss.gated_score == 0.0 and not miss.format_ok


def test_dsbench_deterministic_match_and_evaluator(tmp_path):
    """DSBench: letter/numeric golds grade deterministically; only free-form golds
    reach the LLM judge; gold never appears in prompt or sandbox files."""
    from qea.bench_dsbench import DSTask, DSBenchEvaluator, deterministic_match
    from qea.worker_runtime import WorkerRun

    assert deterministic_match("D", "blah\nFinal answer: D") is True
    assert deterministic_match("D", "Final answer: B") is False
    assert deterministic_match("D", "I think D or maybe B") is None      # ambiguous -> judge
    assert deterministic_match("1661626", "Final answer: 1,661,626") is True
    assert deterministic_match("1661626", "Final answer: $1,661,625") is False
    assert deterministic_match("{'Q': 1}", "Final answer: whatever") is None  # dict -> judge

    calls = []
    class JudgeLLM:
        def complete(self, prompt, *, role="judge", **kw):
            calls.append(prompt)
            return "True"

    t = DSTask(task_id="c_q1", subtype="c", prompt="p", reference_files=[],
               deliverable_exts=[], gold="D", question_text="q?")
    ev = DSBenchEvaluator(JudgeLLM())
    r = ev.evaluate(t, WorkerRun("Final answer: D", [], {}))
    assert r.gated_score == 1.0 and not calls                 # deterministic, no judge call
    r2 = ev.evaluate(t, WorkerRun("maybe D, maybe B", [], {}))
    assert r2.gated_score == 1.0 and len(calls) == 1          # fell back to judge
    assert "The true answer is" in calls[0]                   # official prompt shape


def test_apex_ib_loader_and_evaluator(tmp_path):
    """APEX-IB: rubric conversion to {points, criterion}, world staged via
    vm_setup_cmd (never as local reference uploads), xlsx dump reaches the judge."""
    import openpyxl
    from qea.bench_apex import APEXEvaluator, load_apex_ib
    from qea.worker_runtime import WorkerRun

    tasks = load_apex_ib()
    assert len(tasks) == 160
    assert sum(1 for t in tasks if t.deliverable_exts) == 27      # file deliverables
    t = tasks[0]
    assert t.rubric_items and set(t.rubric_items[0]) == {"points", "criterion"}
    assert "world_files_zipped" in t.vm_setup_cmd and t.subtype in t.vm_setup_cmd
    assert all("gold" not in r.lower() for r in t.reference_files)  # overlays only

    seen = {}
    class StubLLM:
        def complete(self, prompt, *, role="judge", **kw):
            seen["prompt"] = prompt
            return '{"1": true}'

    xl = tmp_path / "model.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws["A1"] = "Net Income"; ws["B1"] = 1772
    wb.save(xl)
    answer_task = next(x for x in tasks if not x.deliverable_exts)
    ev = APEXEvaluator(StubLLM(), k=1)
    r = ev.evaluate(answer_task, WorkerRun("The value is 24.9x", [xl], {}))
    assert r.gated_score > 0
    assert "Net Income" in seen["prompt"] and "1772" in seen["prompt"]  # dump attached
