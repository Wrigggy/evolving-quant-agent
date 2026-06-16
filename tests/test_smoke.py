"""Smoke test: the synthetic plumbing fixture must exhibit the three §5.4
mechanism signals.

Also checks the reference computations and the perturbation-probe logic directly,
so the hard verifier itself is trusted (it is what drives evolution).
"""

import math

import pytest

from qea.loop import Config, acceptance_signals, run_synthetic_fixture
from qea.tasks import (
    bs_call_reference,
    current_ratio_reference,
    load_gdpval_a_pile,
    loan_amort_reference,
    npv_reference,
)
from qea.verifier import HardVerifier


# --------------------------------------------------------------------------- #
# Reference computations.                                                     #
# --------------------------------------------------------------------------- #
def test_black_scholes_atm():
    # S=K=100, r=0.05, sigma=0.2, T=1 -> ~10.4506 (textbook value)
    p = bs_call_reference({"S": 100, "K": 100, "r": 0.05, "sigma": 0.2, "T": 1.0})["call_price"]
    assert abs(p - 10.4506) < 1e-3


def test_amort_fully_amortizes():
    out = loan_amort_reference({"principal": 500_000, "annual_rate": 0.06, "years": 30, "payments_per_year": 12})
    assert abs(out["payment"] - 2997.75) < 0.5     # ~$2,997.75/mo
    assert abs(out["final_balance"]) < 1e-2         # amortizes to ~0


def test_current_ratio():
    out = current_ratio_reference({"current_assets": 1_250_000, "current_liabilities": 800_000})
    assert abs(out["current_ratio"] - 1.5625) < 1e-9
    assert out["working_capital"] == 450_000


def test_npv_irr():
    out = npv_reference({"rate": 0.09, "cashflows": [-100_000_000, 28_000_000, 33_000_000, 39_000_000, 46_000_000]})
    assert out["npv"] > 0                            # positive NPV at 9%
    # IRR is the rate where NPV == 0
    from qea.tasks import _npv
    assert abs(_npv(out["irr"], [-100_000_000, 28_000_000, 33_000_000, 39_000_000, 46_000_000])) < 1.0


# --------------------------------------------------------------------------- #
# Perturbation probe: a hardcoded constant must FAIL the probe (the integrity   #
# guard's whole point). A general solution must PASS.                          #
# --------------------------------------------------------------------------- #
def test_perturbation_probe_kills_hardcoding():
    task = next(t for t in load_gdpval_a_pile() if t.task_id == "A_opt_01")
    hv = HardVerifier()
    base = task.reference(task.inputs)["call_price"]

    hardcoded = f"def solve(inputs):\n    return {{'call_price': {base}}}\n"
    res_hc = hv._score_real(task, hardcoded, k=2)
    assert res_hc.base_pass and not res_hc.probe_pass and not res_hc.oos_pass

    general = (
        "def solve(inputs):\n"
        "    S,K,r,s,T = inputs['S'],inputs['K'],inputs['r'],inputs['sigma'],inputs['T']\n"
        "    d1=(math.log(S/K)+(r+0.5*s*s)*T)/(s*math.sqrt(T)); d2=d1-s*math.sqrt(T)\n"
        "    nc=lambda x:0.5*(1+math.erf(x/math.sqrt(2)))\n"
        "    return {'call_price': S*nc(d1)-K*math.exp(-r*T)*nc(d2)}\n"
    )
    res_g = hv._score_real(task, general, k=2)
    assert res_g.base_pass and res_g.probe_pass and res_g.oos_pass


# --------------------------------------------------------------------------- #
# The three §5.4 mechanism signals (synthetic plumbing fixture).              #
# --------------------------------------------------------------------------- #
@pytest.fixture(scope="module")
def synth_run(tmp_path_factory):
    cfg = Config(mock=True, n_iters=4, k=2, results_dir=str(tmp_path_factory.mktemp("results")))
    return run_synthetic_fixture(cfg)


def test_signal_1_causal_loop(synth_run):
    arm = synth_run
    r1 = arm.records[0]
    # EVAL -> DIAGNOSE -> WORKSPACE -> VERDICT all reference the same root cause
    assert r1.root_cause_tag == "Hardcoding"
    assert r1.verdict == "EFFECTIVE" and r1.kept
    assert "integrity_guard" in arm.final_harness_summary.get("validator", [])


def test_signal_2_monotonic_oos(synth_run):
    traj = synth_run.oos_trajectory
    assert all(traj[i] <= traj[i + 1] for i in range(len(traj) - 1)), traj
    assert max(traj) > min(traj), traj                # strict rise at least once
    assert traj[0] == 0                               # seed hardcodes -> 0 OOS


def test_signal_3_correct_rollback(synth_run):
    recs = synth_run.records
    assert any(r.verdict == "HARMFUL" and not r.kept for r in recs)        # broke code_exec
    assert any(r.verdict == "INEFFECTIVE" and not r.kept for r in recs)    # overfit killed by probe
    assert any(r.blocked for r in recs)                                    # repeat blocked by buffer


def test_acceptance_all_signals(synth_run):
    assert all(acceptance_signals(synth_run).values())


def test_capability_wall_never_solved(synth_run):
    # The amortization subtype includes a capability wall: even the evolved
    # harness cannot lift it (iron law 1), and it stays visible per-subtype.
    n_oos, n_total = synth_run.final_per_subtype["amortization"]
    assert n_oos < n_total


def test_gdpval_rubric_grader_continuous():
    from qea.tasks import BTask, _parse_rubric_json
    from qea.verifier import SoftJudge

    items = _parse_rubric_json('[{"score":2,"criterion":"a"},{"score":1,"criterion":"b"}]')
    assert items == [{"points": 2.0, "criterion": "a"}, {"points": 1.0, "criterion": "b"}]

    class StubLLM:
        def complete(self, prompt, *, role="judge"):
            return 'sure: {"1": true, "2": false, "3": true}'

    t = BTask(task_id="b", subtype="x", prompt="p", rubric="",
              rubric_items=[{"points": 2, "criterion": "a"},
                            {"points": 1, "criterion": "b"},
                            {"points": 1, "criterion": "c"}])
    r = SoftJudge(StubLLM()).score(t, "deliverable", None, mock=False, k=1)
    # earned 2+1=3 of total 4 -> 0.75 (NO quantize to {0,0.5,1})
    assert abs(r.score - 0.75) < 1e-9
    # per-criterion verdicts are exposed on the result
    assert r.criterion_verdicts == {"1": True, "2": False, "3": True}


# --------------------------------------------------------------------------- #
# Regression locks for the review fixes.                                      #
# --------------------------------------------------------------------------- #
def test_irr_tolerance_is_per_metric():
    # A solution with correct NPV but a wildly wrong IRR must FAIL (the old bug:
    # tol=1.0 applied to a ~0.15 IRR let ±643% error pass).
    task = next(t for t in load_gdpval_a_pile() if t.task_id == "A_val_01")
    hv = HardVerifier()
    exp = task.reference(task.inputs)
    bad_irr = (
        "def solve(inputs):\n"
        f"    cfs=inputs['cashflows']; r=inputs['rate']\n"
        "    npv=sum(cf/(1+r)**t for t,cf in enumerate(cfs))\n"
        f"    return {{'npv': npv, 'irr': 0.99}}\n"   # NPV right, IRR absurd
    )
    res = hv._score_real(task, bad_irr, k=2)
    assert not res.base_pass  # wrong IRR must fail the base check now


def test_unattributed_regression_downgrades_verdict():
    from qea.falsify import evaluate_changes
    from qea.harness import Edit
    # predicts to fix A, B; both flip, but unrelated C regresses -> not EFFECTIVE.
    edit = Edit(op="add", slot="validator", component_name="x", predicted_fixes=["A", "B"], risk_tasks=[])
    ev = evaluate_changes(edit, {"flipped": ["A", "B"], "regressed": ["C"]})
    assert ev["verdict"] in ("MIXED", "HARMFUL")
    assert ev["unattributed_regressions"] == ["C"]


def test_sandbox_allows_safe_imports_blocks_others():
    from qea.verifier import safe_exec_solve
    # models naturally write `import math` — must work
    ok = "def solve(inputs):\n    import math\n    return {'x': math.sqrt(inputs['v'])}\n"
    assert safe_exec_solve(ok, {"v": 16.0})["x"] == 4.0
    # disallowed imports stay blocked
    bad = "def solve(inputs):\n    import os\n    return {'x': 1}\n"
    with pytest.raises(Exception):
        safe_exec_solve(bad, {})


def test_soft_gate_noise_aware():
    from qea.falsify import decide_keep_soft
    assert decide_keep_soft(0.618, 0.651, 0.02) is True    # gain 0.033 > noise 0.02 -> keep
    assert decide_keep_soft(0.618, 0.651, 0.05) is False   # gain within noise -> reject
    assert decide_keep_soft(0.618, 0.560, 0.02) is False   # regression -> reject


def test_signature_no_collision_past_120_chars():
    from qea.harness import Edit
    base = "x" * 130
    a = Edit(op="add", slot="prompt", component_name="p", content=base + "AAA")
    b = Edit(op="add", slot="prompt", component_name="p", content=base + "BBB")
    assert a.signature() != b.signature()


def test_gdpval_soft_mock_pct_gate(tmp_path):
    # End-to-end mock: iter-1 integrity_guard lifts the mock soft score
    # (0.45 -> 0.72), beating the noise floor, so the % gate keeps it and the
    # mean-score trajectory rises.
    from qea.loop import Config, run_gdpval_soft
    res = run_gdpval_soft(Config(mock=True, n_iters=2, k=2, results_dir=str(tmp_path)))
    assert res.n_kept >= 1
    assert res.mean_score_trajectory[-1] > res.mean_score_trajectory[0]
    assert not hasattr(res, "final_elo_vs_seed")  # pairwise fields gone


def test_provider_pin_is_per_model_official(monkeypatch):
    # Every model must pin ITS OWN official provider; no cross-pinning.
    from qea.llm import provider_for, resolve_provider_map
    monkeypatch.setenv("QEA_PROVIDER_ORDER", "deepseek")
    pmap = resolve_provider_map()
    assert provider_for("deepseek/deepseek-v4-pro", pmap) == "deepseek"
    assert provider_for("qwen/qwen3.7-max", pmap) == "alibaba"      # built-in official map
    assert provider_for("google/gemini-3.1-pro-preview", pmap) is None  # unpinned -> free routing
    monkeypatch.setenv("QEA_PROVIDER_MAP", "google=google-ai-studio,qwen=other")
    pmap = resolve_provider_map()
    assert provider_for("google/gemini-3.1-pro-preview", pmap) == "google-ai-studio"
    assert provider_for("qwen/qwen3.7-max", pmap) == "other"        # env overrides built-in


def test_gdpval_local_fork_preferred():
    # The rubric fork in data/gdpval/ must be used (no network) when present.
    from qea.tasks import _GDPVAL_LOCAL_PARQUET, _load_gdpval_df
    if not _GDPVAL_LOCAL_PARQUET.exists():
        pytest.skip("local GDPval fork not present (run scripts/fork_gdpval.py)")
    df, src = _load_gdpval_df()
    assert src == "local fork"
    assert len(df) == 220
    assert {"task_id", "rubric_json", "rubric_pretty"} <= set(df.columns)


def test_deliverable_cache_stable_per_harness():
    # The same harness must yield the SAME deliverable within a run (cache hit),
    # so re-evaluating an unchanged harness does not wobble from regeneration.
    from qea.harness import seed_harness, Edit
    from qea.loop import _DeliverableCache
    calls = {"n": 0}

    def gen():
        calls["n"] += 1
        return f"deliverable v{calls['n']}"

    cache = _DeliverableCache()
    h = seed_harness()
    a = cache.get_or_make("t1", h, gen)
    b = cache.get_or_make("t1", h, gen)
    assert a == b and calls["n"] == 1            # second call is a cache hit
    # a different harness must miss the cache and regenerate
    h2 = h.clone(); h2.apply(Edit(op="add", slot="memory", component_name="kb", content="x"))
    c = cache.get_or_make("t1", h2, gen)
    assert c == "deliverable v2" and calls["n"] == 2


def test_b_debugger_attributes_and_firewalls():
    from qea.tasks import BTask
    from qea.verifier import TaskResult
    from qea.falsify import EvalSummary
    from qea.debugger import diagnose_b_pile

    class CriticLLM:
        def complete(self, prompt, *, role="judge"):
            if "Classify" in prompt:
                return '{"root_cause_tag": "MissingDomainKnowledge", "target_slot": "memory"}'
            return "The deliverable omits the going-concern analysis the rubric requires."

    res = {"t1": TaskResult("t1", "Accountants and Auditors", "B", False, False, False, 0.3, 0.0,
                            None, criterion_verdicts={"1": True, "2": False})}
    tasks = [BTask(task_id="t1", subtype="Accountants and Auditors", prompt="audit memo", rubric="",
                   rubric_items=[{"points": 1, "criterion": "states ratios"},
                                 {"points": 2, "criterion": "flags going-concern triggers"}],
                   gold="SECRET-ANSWER-12345")]
    diag = diagnose_b_pile(EvalSummary(res, {"t1": "weak memo"}), tasks, llm=CriticLLM(), mode="hybrid")
    assert diag.root_cause_tag == "MissingDomainKnowledge"
    assert diag.suggested_target_slot == "memory"
    payload = diag.proposer_payload()
    blob = repr(payload)
    assert "SECRET-ANSWER-12345" not in blob
    assert "going-concern triggers" not in blob
    assert "t1" in payload["predicted_fix_task_ids"]


def test_b_debugger_survives_llm_outage():
    # A sustained judge/critic outage (llm.complete raises after retries) must NOT
    # crash the run — diagnose_b_pile degrades to a fallback note + default tag,
    # mirroring evaluate()'s per-task fault tolerance. (Regression: a real 10-iter
    # run died here when the critic call propagated an uncaught RuntimeError.)
    from qea.tasks import BTask
    from qea.verifier import TaskResult
    from qea.falsify import EvalSummary
    from qea.debugger import diagnose_b_pile

    class BoomLLM:
        def complete(self, prompt, *, role="judge"):
            raise RuntimeError("LLM failed after 5 retries: Connection error.")

    res = {"t1": TaskResult("t1", "Accountants and Auditors", "B", False, False, False, 0.3, 0.0,
                            None, criterion_verdicts={"1": False})}
    tasks = [BTask(task_id="t1", subtype="Accountants and Auditors", prompt="memo", rubric="",
                   rubric_items=[{"points": 1, "criterion": "x"}], gold="SECRET")]
    diag = diagnose_b_pile(EvalSummary(res, {"t1": "weak"}), tasks, llm=BoomLLM(), mode="hybrid")
    assert diag.root_cause_tag == "WrongStructure"           # default tag, no crash
    assert "t1" in diag.predicted_fix_task_ids
    assert "SECRET" not in repr(diag.proposer_payload())     # firewall holds on the fallback path


def test_propose_real_prompt_has_no_answers():
    from qea.agents import _propose_real
    from qea.harness import seed_harness
    from qea.falsify import RejectedEditBuffer, EvalSummary
    from qea.verifier import TaskResult

    captured = {}
    class LLM:
        def complete(self, prompt, *, role="agent"):
            captured["p"] = prompt
            return ('{"slot":"memory","component_name":"kb","content":"general finance knowledge",'
                    '"summary":"add kb","predicted_fixes":["t1"],"risk_tasks":[]}')
    diag = {"root_cause_tag": "MissingDomainKnowledge", "deficiency_category": "1 task",
            "suggested_target_slot": "memory", "predicted_fix_task_ids": ["t1"],
            "overview": "MissingDomainKnowledge deficiency", "_b_pile": True}
    # Populate the eval with sentinels in EVERY field the B-branch must NOT read
    # (result subtype/error + the deliverable text). A firewall regression that
    # reaches into eval_summary would leak one of these into the prompt.
    res = {"t1": TaskResult("t1", "SECRET-SUBTYPE", "B", False, False, False, 0.3, 0.0,
                            "SECRET-ERROR", criterion_verdicts={"1": False})}
    evalsum = EvalSummary(res, {"t1": "SECRET-DELIVERABLE going-concern analysis"})
    _propose_real(1, evalsum, diag, seed_harness(), RejectedEditBuffer(), LLM())
    assert "SECRET" not in captured["p"]   # no ground truth from eval_summary leaked
    assert "MissingDomainKnowledge" in captured["p"]  # sanitized signal present


def test_leakage_guard_blocks_copied_answer():
    from qea.verifier import LeakageGuard
    from qea.harness import Edit
    corpus = ["flags any going-concern triggers in the liquidity position"]
    guard = LeakageGuard(corpus, threshold=0.6)
    leak = Edit(op="add", slot="memory", component_name="kb",
                content="Always flags any going-concern triggers in the liquidity position.")
    assert guard.is_leak(leak) is True
    ok = Edit(op="add", slot="prompt", component_name="p",
              content="Structure the memo with a clear recommendation section.")
    assert guard.is_leak(ok) is False


def test_rubric_corpus_collects_criteria():
    from qea.tasks import BTask, rubric_corpus
    tasks = [BTask(task_id="t", subtype="x", prompt="p", rubric="overall rubric text",
                   rubric_items=[{"points": 1, "criterion": "states ratios"},
                                 {"points": 2, "criterion": "flags going-concern"}])]
    corpus = rubric_corpus(tasks)
    assert "states ratios" in corpus and "flags going-concern" in corpus
    assert "overall rubric text" in corpus


def test_benchmark_owns_grader_and_corpus():
    from qea.benchmark import gdpval_benchmark
    bm = gdpval_benchmark(broad=False, allow_download=False)  # offline fixtures
    assert bm.name == "gdpval_finance"
    assert bm.tasks and all(t.pile == "B" for t in bm.tasks)
    assert bm.grader is not None
    assert isinstance(bm.answer_corpus, list) and len(bm.answer_corpus) > 0
    assert bm.debugger_kind == "b_pile"


def test_exec_artifact_produces_xlsx_and_scrubs_env(tmp_path, monkeypatch):
    import pytest
    pytest.importorskip("openpyxl")
    from qea.sandbox import exec_artifact
    monkeypatch.setenv("OPENROUTER_API_KEY", "SECRET_KEY_DO_NOT_LEAK")
    code = (
        "import openpyxl, os\n"
        "assert 'OPENROUTER_API_KEY' not in os.environ, 'secret leaked into child'\n"
        "wb = openpyxl.Workbook(); ws = wb.active; ws['A1'] = 'hello'\n"
        "wb.save('report.xlsx')\n"
    )
    res = exec_artifact(code, timeout=10.0)
    assert res.status == "success"
    assert len(res.paths) == 1 and res.paths[0].name == "report.xlsx"


def test_exec_artifact_error_and_timeout_dont_crash_parent():
    import pytest
    from qea.sandbox import exec_artifact
    err = exec_artifact("raise RuntimeError('boom')\n", timeout=10.0)
    assert err.status == "error" and "boom" in err.stderr
    slow = exec_artifact("while True:\n    pass\n", timeout=1.0)
    assert slow.status == "timeout"   # parent process is still alive to assert this


def test_render_xlsx_dumps_values_and_formulas(tmp_path):
    import pytest
    pytest.importorskip("openpyxl")
    import openpyxl
    from qea.artifacts import render_xlsx
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Summary"
    ws["A1"] = "Revenue"; ws["B1"] = 1000; ws["B2"] = "=B1*2"
    p = tmp_path / "model.xlsx"; wb.save(p)
    text = render_xlsx(p)
    assert "[ARTIFACT FILE: model.xlsx]" in text
    assert 'Sheet "Summary"' in text
    assert "'Revenue'" in text and "1000" in text
    assert "=B1*2" in text          # formula string is rendered (data_only=False)


def test_extract_openpyxl_code():
    from qea.artifacts import extract_openpyxl_code
    md = "Here is the workbook:\n```python\nimport openpyxl\nwb=openpyxl.Workbook()\nwb.save('x.xlsx')\n```\nDone."
    code = extract_openpyxl_code(md)
    assert code is not None and "openpyxl" in code and "save(" in code
    assert extract_openpyxl_code("just a plain memo, no code") is None
    assert extract_openpyxl_code("```python\nprint('hi')\n```") is None  # not an artifact block
    # robust to fence-tag variation (```py / ```Python) and whitespace in .save(
    assert extract_openpyxl_code("```py\nimport openpyxl\nopenpyxl.Workbook().save( 'x.xlsx' )\n```") is not None
    assert extract_openpyxl_code("```Python\nimport openpyxl\nopenpyxl.Workbook().save('x.xlsx')\n```") is not None


def _BTaskStub(tid="b1"):
    from qea.tasks import BTask
    return BTask(task_id=tid, subtype="Accountants and Auditors", prompt="produce report.xlsx", rubric="")

def test_assemble_artifact_deliverable_produces_and_persists(tmp_path):
    import pytest
    pytest.importorskip("openpyxl")
    from qea.artifacts import assemble_artifact_deliverable
    llm_text = ("Here is the workbook.\n```python\nimport openpyxl\n"
                "wb=openpyxl.Workbook(); ws=wb.active; ws.title='Summary'; ws['A1']='Total'; ws['B1']=42\n"
                "wb.save('report.xlsx')\n```")
    artifact_dir = tmp_path / "artifacts"
    out = assemble_artifact_deliverable(llm_text, _BTaskStub(), artifact_dir)
    assert "[ARTIFACT FILE: report.xlsx]" in out and "'Total'" in out
    assert "import openpyxl" not in out            # raw code stripped from the narrative
    assert (artifact_dir / "b1" / "report.xlsx").exists()   # persisted under <dir>/<task_id>/

def test_assemble_artifact_deliverable_text_and_error_paths(tmp_path):
    from qea.artifacts import assemble_artifact_deliverable
    # plain text task -> unchanged
    assert assemble_artifact_deliverable("just a memo", _BTaskStub(), tmp_path) == "just a memo"
    # erroring artifact code -> graceful: keep the narrative, no crash
    bad = "```python\nimport openpyxl\nraise RuntimeError('x')\nwb=1\n.save('y.xlsx')\n```narrative"
    out = assemble_artifact_deliverable(bad, _BTaskStub(), tmp_path)
    assert "[ARTIFACT FILE" not in out             # nothing produced

def test_assemble_artifact_deliverable_survives_corrupt_xlsx(tmp_path):
    import pytest
    pytest.importorskip("openpyxl")
    from qea.artifacts import assemble_artifact_deliverable
    # child exits 0 but writes a NON-xlsx file named report.xlsx -> exec "success" but
    # render_xlsx would raise BadZipFile. Must degrade to a placeholder, not crash.
    bad = ("```python\nimport openpyxl\n# not a real .save( call:\n"
           "open('report.xlsx','w').write('not a zip')\n```done")
    out = assemble_artifact_deliverable(bad, _BTaskStub(), tmp_path)
    assert "report.xlsx" in out and "unreadable workbook" in out   # placeholder, no raise
